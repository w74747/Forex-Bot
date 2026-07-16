"""
flask_dashboard.py
"""

import os
import logging
from datetime import datetime
from io import BytesIO

import psycopg2
from flask import Flask, render_template, jsonify, request, send_file
from psycopg2.extras import RealDictCursor

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
DATABASE_URL = os.environ.get("DATABASE_URL")
DEBUG = os.environ.get("FLASK_DEBUG", "false").lower() == "true"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("dashboard")

def get_db():
    try:
        return psycopg2.connect(DATABASE_URL, connect_timeout=5)
    except Exception as e:
        logger.error(f"[DB] {e}")
        return None

def query_trades(status=None, symbol=None, start_date=None, end_date=None, limit=100):
    conn = get_db()
    if not conn:
        return []
    
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            query = "SELECT * FROM live_paper_trades WHERE 1=1"
            params = []
            
            if status:
                query += " AND status = %s"
                params.append(status)
            if symbol:
                query += " AND symbol = %s"
                params.append(symbol)
            if start_date:
                query += " AND DATE(opened_at) >= %s"
                params.append(start_date)
            if end_date:
                query += " AND DATE(opened_at) <= %s"
                params.append(end_date)
            
            query += " ORDER BY opened_at DESC LIMIT %s"
            params.append(limit)
            
            cur.execute(query, params)
            return [dict(row) for row in cur.fetchall()]
    except Exception as e:
        logger.error(f"[DB] {e}")
        return []
    finally:
        conn.close()

def get_statistics():
    conn = get_db()
    if not conn:
        return {}
    
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT 
                    COUNT(*) as total_trades,
                    SUM(CASE WHEN net_pnl > 0 THEN 1 ELSE 0 END) as winning_trades,
                    SUM(CASE WHEN net_pnl < 0 THEN 1 ELSE 0 END) as losing_trades,
                    SUM(net_pnl) as total_pnl,
                    AVG(net_pnl) as avg_pnl,
                    MAX(net_pnl) as max_win,
                    MIN(net_pnl) as max_loss
                FROM live_paper_trades 
                WHERE status = 'CLOSED'
            """)
            stats = dict(cur.fetchone() or {})
            
            total = stats.get('total_trades') or 0
            if total > 0:
                win_rate = (stats.get('winning_trades') or 0) / total * 100
                stats['win_rate'] = round(win_rate, 2)
            else:
                stats['win_rate'] = 0
            
            cur.execute("SELECT COUNT(*) as open_positions FROM live_paper_trades WHERE status = 'OPEN'")
            open_stats = dict(cur.fetchone() or {})
            stats['open_positions'] = open_stats.get('open_positions', 0)
            
            return stats
    except Exception as e:
        logger.error(f"[DB] {e}")
        return {}
    finally:
        conn.close()

def get_daily_equity_curve(days=7):
    conn = get_db()
    if not conn:
        return []
    
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"""
                WITH daily_pnl AS (
                    SELECT 
                        DATE(closed_at) as trade_date,
                        SUM(net_pnl) as daily_sum
                    FROM live_paper_trades
                    WHERE status = 'CLOSED'
                    AND closed_at >= NOW() - INTERVAL '{days} days'
                    GROUP BY DATE(closed_at)
                    ORDER BY trade_date
                )
                SELECT 
                    trade_date,
                    daily_sum,
                    SUM(daily_sum) OVER (ORDER BY trade_date) as cumulative_pnl
                FROM daily_pnl
            """)
            return [dict(row) for row in cur.fetchall()]
    except Exception as e:
        logger.error(f"[DB] {e}")
        return []
    finally:
        conn.close()

def get_logs(log_type=None, start_date=None, end_date=None, limit=200):
    conn = get_db()
    if not conn:
        return []
    
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            query = "SELECT * FROM system_logs WHERE 1=1"
            params = []
            
            if log_type:
                query += " AND log_type = %s"
                params.append(log_type)
            if start_date:
                query += " AND DATE(created_at) >= %s"
                params.append(start_date)
            if end_date:
                query += " AND DATE(created_at) <= %s"
                params.append(end_date)
            
            query += " ORDER BY created_at DESC LIMIT %s"
            params.append(limit)
            
            cur.execute(query, params)
            return [dict(row) for row in cur.fetchall()]
    except Exception as e:
        logger.error(f"[DB] {e}")
        return []
    finally:
        conn.close()

@app.route('/')
def index():
    return render_template('dashboard.html')

@app.route('/api/trades', methods=['GET'])
def api_trades():
    status = request.args.get('status')
    symbol = request.args.get('symbol')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    limit = int(request.args.get('limit', 100))
    
    trades = query_trades(status, symbol, start_date, end_date, limit)
    for trade in trades:
        if trade.get('opened_at'):
            trade['opened_at'] = trade['opened_at'].isoformat()
        if trade.get('closed_at'):
            trade['closed_at'] = trade['closed_at'].isoformat()
    
    return jsonify(trades)

@app.route('/api/statistics', methods=['GET'])
def api_statistics():
    stats = get_statistics()
    return jsonify(stats)

@app.route('/api/equity-curve', methods=['GET'])
def api_equity_curve():
    days = int(request.args.get('days', 7))
    curve = get_daily_equity_curve(days)
    for point in curve:
        if point.get('trade_date'):
            point['trade_date'] = point['trade_date'].isoformat()
    return jsonify(curve)

@app.route('/api/logs', methods=['GET'])
def api_logs():
    log_type = request.args.get('log_type')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    limit = int(request.args.get('limit', 200))
    
    logs = get_logs(log_type, start_date, end_date, limit)
    for log in logs:
        if log.get('created_at'):
            log['created_at'] = log['created_at'].isoformat()
    
    return jsonify(logs)

@app.route('/api/export/trades', methods=['GET'])
def export_trades_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 500
    
    status = request.args.get('status')
    symbol = request.args.get('symbol')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    trades = query_trades(status, symbol, start_date, end_date, limit=10000)
    if not trades:
        return jsonify({"error": "No data"}), 400
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    
    headers = ["ID", "Symbol", "Direction", "Entry", "SL", "TP", "Status", "Exit", "Reason", "Gross PnL", "Commission", "Net PnL", "Open Time", "Close Time"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, trade in enumerate(trades, 2):
        data = [
            trade.get('id'),
            trade.get('symbol'),
            trade.get('direction'),
            f"{trade.get('entry_price', 0):.5f}",
            f"{trade.get('sl_price', 0):.5f}",
            f"{trade.get('tp_price', 0):.5f}",
            trade.get('status'),
            f"{trade.get('exit_price', 0):.5f}" if trade.get('exit_price') else "",
            trade.get('exit_reason', ''),
            f"{trade.get('gross_pnl', 0):.2f}",
            f"{trade.get('commission', 0):.2f}",
            f"{trade.get('net_pnl', 0):.2f}",
            str(trade.get('opened_at', '')),
            str(trade.get('closed_at', '')) if trade.get('closed_at') else "",
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 12 and trade.get('net_pnl', 0) > 0:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif col == 12 and trade.get('net_pnl', 0) < 0:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    ws.column_dimensions['A'].width = 8
    for col in 'BCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'trades_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@app.route('/api/export/logs', methods=['GET'])
def export_logs_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 500
    
    log_type = request.args.get('log_type')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    logs = get_logs(log_type, start_date, end_date, limit=10000)
    if not logs:
        return jsonify({"error": "No data"}), 400
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    
    headers = ["ID", "Type", "Message", "Created At"]
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, log in enumerate(logs, 2):
        data = [log.get('id'), log.get('log_type'), log.get('message'), str(log.get('created_at', ''))]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@app.errorhandler(404)
def not_found(error):
    return jsonify({"error": "Not found"}), 404

@app.errorhandler(500)
def server_error(error):
    logger.error(f"Error: {error}")
    return jsonify({"error": "Server error"}), 500

if __name__ == '__main__':
    logger.info("Starting Flask Dashboard...")
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=DEBUG, use_reloader=False)
